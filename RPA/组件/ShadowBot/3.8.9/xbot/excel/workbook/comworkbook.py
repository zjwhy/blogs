import openpyxl
import typing
import os
import atexit

from ..worksheet.baseworksheet import BaseWorkSheet
from ..worksheet.comworksheet import ComWorkSheet
from .baseworkbook import BaseWorkBook

_opened_workbooks = []

def _close_workbooks_atexit():
    global _opened_workbooks
    for workbook in _opened_workbooks:
        if workbook.launch_way == 'get_active':
            workbook.workbook.Save()
            workbook.xlApp.DisplayAlerts = True
        else:
            if workbook.xlApp.Visible:  
                workbook.xlApp.DisplayAlerts = True
            else:
                workbook.workbook.Save()
                workbook.workbook.Close()
                if workbook.xlApp.Workbooks.Count == 0:
                    workbook.xlApp.Quit()
                
    _opened_workbooks = []

atexit.register(_close_workbooks_atexit)


class ComWorkBook(BaseWorkBook):
    def __init__(self, workbook, launch_way, original_file, xlApp):
        self.workbook = workbook
        self.launch_way = launch_way
        self.original_file = original_file
        self.xlApp = xlApp

        global _opened_workbooks
        _opened_workbooks.append(self)


    def save(self) -> typing.NoReturn:
        if self.launch_way == 'create':
            self.workbook.SaveAs(self.original_file)
        else:
            self.workbook.Save()
            

    def save_as(self, filename) -> typing.NoReturn:
        self.workbook.SaveAs(filename)


    def close(self) -> typing.NoReturn:
        self.workbook.Close()

        if self.xlApp.Workbooks.Count == 0:
            self.xlApp.Quit()

        global _opened_workbooks
        _opened_workbooks.remove(self)


    def create_sheet(self, name, create_way) -> BaseWorkSheet:
        # 1、检查新增sheet页的名称
        if name in [ws.Name for ws in self.workbook.Worksheets]:
            raise ValueError(f'已存在名为{name}的sheet页')
        # 2、创建sheet页
        if create_way == 'first':
            sheet = self.workbook.Worksheets.Add(self.workbook.Worksheets(1))
            sheet.Name = name
            return ComWorkSheet(sheet)
        else:
            totalSheetCount = self.workbook.Worksheets.Count
            sheet = self.workbook.Worksheets.Add(None, self.workbook.Worksheets(totalSheetCount))
            sheet.Name = name
            return ComWorkSheet(sheet)
    

    def active_sheet_by_index(self, index) -> typing.NoReturn:
        self.workbook.Worksheets(index).activate


    def active_sheet_by_name(self, name) -> typing.NoReturn:
        self.workbook.Worksheets(name).activate
    

    def get_sheet_by_index(self, index) -> BaseWorkSheet:
        return ComWorkSheet(self.workbook.Worksheets(index))


    def get_sheet_by_name(self, name) -> BaseWorkSheet:
        return ComWorkSheet(self.workbook.Worksheets(name))
    

    def get_active_sheet(self) -> BaseWorkSheet:
        return ComWorkSheet(self.workbook.ActiveSheet)

    
    def execute_macro(self, macro) -> typing.NoReturn:
        pass


    def get_all_sheets(self) -> typing.List[BaseWorkSheet]:
        return [ComWorkSheet(ws) for ws in self.workbook.Worksheets]


    def delete_sheet(self, name) -> typing.NoReturn:
        if name not in [ws.Name for ws in self.workbook.Worksheets]:
            raise ValueError(f'不存在名为{name}的sheet页')
        self.workbook.Worksheets(name).Delete()


    def copy_sheet(self, name, new_name) -> BaseWorkSheet:
        if name not in [ws.Name for ws in self.workbook.Worksheets]:
            raise ValueError(f'不存在名为{name}的sheet页')
        if new_name in [ws.Name for ws in self.workbook.Worksheets]:
            raise ValueError(f'已存在名为{new_name}的sheet页')

        count = self.workbook.Worksheets.Count
        self.workbook.Worksheets(name).Copy(None, self.workbook.Worksheets(count))

        sheet = self.workbook.Worksheets(count+1)
        sheet.name = new_name

        return ComWorkSheet(sheet)


    def copy_sheet_to_workbook(self, name, workbook, new_name) -> BaseWorkSheet:
        if name not in [ws.Name for ws in self.workbook.Worksheets]:
            raise ValueError(f'不存在名为{name}的sheet页')
        if new_name in [ws.Name for ws in workbook.workbook.Worksheets]:
            raise ValueError(f'已存在名为{new_name}的sheet页')

        count = workbook.workbook.Worksheets.Count
        self.workbook.Worksheets(name).Copy(None, workbook.workbook.Worksheets(count))

        sheet = workbook.workbook.Worksheets(count+1)
        sheet.name = new_name

        return ComWorkSheet(sheet)


    def get_selected_range(self) -> typing.Tuple:
        selection = self.xlApp.Selection

        if selection.Rows.Count == 0:
            return None

        begin_row_num = selection.Row
        begin_column_name = openpyxl.utils.cell.get_column_letter(selection.Column)
        end_row_num = selection.Row + selection.Rows.Count - 1
        end_column_name = openpyxl.utils.cell.get_column_letter(selection.Column + selection.Columns.Count - 1)

        return (begin_row_num, begin_column_name, end_row_num, end_column_name)