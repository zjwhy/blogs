import openpyxl
import typing

from ..worksheet.baseworksheet import BaseWorkSheet
from ..worksheet.openpyxlworksheet import OpenPyxlWorkSheet
from .baseworkbook import BaseWorkBook

class OpenPyxlWorkBook(BaseWorkBook):
    def __init__(self, workbook, launch_way, original_file):
        self.workbook = workbook
        self.launch_way = launch_way
        self.original_file = original_file
    

    def save(self) -> typing.NoReturn:
        self.workbook.save(self.original_file)


    def save_as(self, filename) -> typing.NoReturn:
        self.workbook.save(filename)


    def close(self) -> typing.NoReturn:
        raise ValueError(f'openpyxl目前不支持关闭操作')


    def create_sheet(self, name, create_way) -> BaseWorkSheet:
        # 1、检查新增sheet页的名称
        if name in self.workbook.sheetnames:
            raise ValueError(f'已存在名为{name}的sheet页')
        # 2、创建sheet页
        if create_way == 'first':
            worksheet = self.workbook.create_sheet(title=name, index=0)
            return OpenPyxlWorkSheet(worksheet)
        else:
            worksheet = self.workbook.create_sheet(title=name)
            return OpenPyxlWorkSheet(worksheet)
    

    def active_sheet_by_index(self, index) -> typing.NoReturn:
        self.workbook.active = index - 1


    def active_sheet_by_name(self, name) -> typing.NoReturn:
        self.workbook.active = self.workbook[name]


    def get_sheet_by_index(self, index) -> BaseWorkSheet:
        sheet_name = self.workbook.sheetnames[index - 1]
        return OpenPyxlWorkSheet(self.workbook[sheet_name])
    

    def get_sheet_by_name(self, name) -> BaseWorkSheet:
        return OpenPyxlWorkSheet(self.workbook[name])
    

    def get_active_sheet(self) -> BaseWorkSheet:
        return OpenPyxlWorkSheet(self.workbook.active)


    def execute_macro(self, macro) -> typing.NoReturn:
        raise ValueError("openpyxl目前不支持运行宏")


    def get_all_sheets(self) -> typing.List[BaseWorkSheet]:
        return [OpenPyxlWorkSheet(ws) for ws in self.workbook.worksheets]


    def delete_sheet(self, name) -> typing.NoReturn:
        if name not in self.workbook.sheetnames:
            raise ValueError(f'不存在名为{name}的sheet页')
        self.workbook.remove(self.workbook[name])


    def copy_sheet(self, name, new_name) -> BaseWorkSheet:
        if name not in self.workbook.sheetnames:
            raise ValueError(f'不存在名为{name}的sheet页')
        if new_name in self.workbook.sheetnames:
            raise ValueError(f'已存在名为{new_name}的sheet页')

        new_sheet = self.workbook.copy_worksheet(self.workbook[name])
        new_sheet.title = new_name

        return OpenPyxlWorkSheet(new_sheet)


    def copy_sheet_to_workbook(self, name, workbook, new_name) -> BaseWorkSheet:
        if name not in self.workbook.sheetnames:
            raise ValueError(f'不存在名为{name}的sheet页')
        if new_name in workbook.workbook.sheetnames:
            raise ValueError(f'已存在名为{new_name}的sheet页')

        dest_worksheet = workbook.create_sheet(new_name, 'last')
        source_worksheet = self.get_sheet_by_name(name)
        values = source_worksheet.get_range(
            1,
            'A',
            source_worksheet.get_row_count(),
            openpyxl.utils.cell.get_column_letter(source_worksheet.get_column_count())
        )
        
        dest_worksheet.set_range(1, 'A', values)

        return OpenPyxlWorkSheet(new_sheet)


    def get_selected_range(self) -> typing.Tuple:
        raise ValueError("openpyxl目前不支持获取选中区域")