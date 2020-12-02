import openpyxl
import typing

from .baseworksheet import BaseWorkSheet

class OpenPyxlWorkSheet(BaseWorkSheet):
    copyed_cells = []

    def __init__(self, worksheet):
        self.worksheet = worksheet
    
    def get_cell(self, row_num, col_name) -> any:
        return self.worksheet[f'{col_name}{row_num}'].value

    
    def get_row(self, row_num) -> typing.List[any]:
        if row_num > self.get_row_count():
            return []

        for row in self.worksheet.iter_rows(min_row=row_num, max_row=row_num, values_only=False):
            row_data = [col.value for col in row]
        return row_data


    def get_column(self, col_name) -> typing.List[any]:
        column_index = openpyxl.utils.cell.column_index_from_string(col_name)
        if column_index > self.get_column_count():
            return []

        begin_row_num = 1
        end_row_num = self.get_row_count()
        # Range的效率和其实际包含的数据无关，和调用Range的次数有关，所以优化成下面这种方式提高效率
        table_data = self.worksheet[f'{col_name}{begin_row_num}:{col_name}{end_row_num}']
        return [row[0].value for row in table_data]


    def get_range(self, begin_row_num, begin_column_name, end_row_num, end_column_name) -> typing.List[typing.List[any]]:
        table_data = []
        for row in self.worksheet[f'{begin_column_name}{begin_row_num}':f'{end_column_name}{end_row_num}']:
            table_data.append([col.value for col in row])

        return table_data


    def set_cell(self, row_num, col_name, value) -> typing.NoReturn:
       self.worksheet[f'{col_name}{row_num}'] = value


    def set_row(self, row_num, values, begin_column_name='A') -> typing.NoReturn:
        # 2.1 将单个值转成[]
        if not isinstance(values, (list, tuple)):    #单个值的情况
            values = [values]
        # 2.2 计算插入点并插入内容
        begin_column_index = openpyxl.utils.cell.column_index_from_string(begin_column_name)
        for col_index, val in enumerate(values, start = begin_column_index):
           self.worksheet.cell(row=row_num, column=col_index).value = val


    def set_column(self, col_name, values, begin_row_num = 1) -> typing.NoReturn:
        # 2.1 将单个值转成[]
        if not isinstance(values, (list, tuple)):
            values = [values]
        # 2.2 写入内容
        col_name_index = openpyxl.utils.cell.column_index_from_string(col_name)
        for row_index, val in enumerate(values, start = begin_row_num):
           self.worksheet.cell(row=row_index, column=col_name_index).value = val


    def set_range(self, row_num, col_name, values) -> typing.NoReturn:
        #2.1 转成[[]]
        if not isinstance(values, (list, tuple)):    #单个值的情况 ()
            values = [[values]]
        else:
            if len(values) == 0:           # []的情况
                values = [[]]
            else:
                if not isinstance(values[0], (list, tuple)):   # [1,2,3,4,5]的情况 (一维)
                    values = [values]
        
        #2.2 写入数据
        column_index = openpyxl.utils.cell.column_index_from_string(col_name)
        for row_value in values:
            for col_index, col_val in enumerate(row_value, start = column_index):
               self.worksheet.cell(row=row_num, column=col_index).value = col_val
            row_num += 1


    def append_row(self, values, begin_column_name = 'A') -> typing.NoReturn:
        # 2.1 将单个值转成[]
        if not isinstance(values, (list, tuple)):    #单个值的情况
            values = [values]
        
        # 2.2 计算插入点并插入内容
        row_num = self.get_row_count() + 1
        begin_column_index = openpyxl.utils.cell.column_index_from_string(begin_column_name)
        for col_index, val in enumerate(values, start = begin_column_index):
           self.worksheet.cell(row=row_num, column=col_index).value = val


    def insert_row(self, row_num, values, begin_column_name = 'A') -> typing.NoReturn:
        # 2.1 将单个值转成[]
        if not isinstance(values, (list, tuple)):    #单个值的情况
            values = [values]
        
        # 2.2 计算插入点并插入内容
        self.worksheet.insert_rows(row_num)
        begin_column_index = openpyxl.utils.cell.column_index_from_string(begin_column_name)
        for col_index, val in enumerate(values, start = begin_column_index):
           self.worksheet.cell(row=row_num, column=col_index).value = val


    def remove_row(self, row_num) -> typing.NoReturn:
       self.worksheet.delete_rows(row_num)


    def clear(self) -> typing.NoReturn:
        total_rows_count = self.get_row_count()
        if total_rows_count > 0:
           self.worksheet.delete_rows(1, total_rows_count)


    def get_row_count(self) -> int:
        if self.worksheet.dimensions == 'A1:A1' and self.worksheet.cell(1, 1).value is None:
            return 0
        else:
            return self.worksheet.max_row


    def get_column_count(self) -> int:
        if self.worksheet.dimensions == 'A1:A1' and self.worksheet.cell(1, 1).value is None:
            return 0
        else:
            return self.worksheet.max_column
    

    def select_range(self, begin_row_num, begin_column_name, end_row_num, end_column_name) -> typing.NoReturn:
        raise ValueError("openpyxl目前不支持选中区域")


    def select_rows(self, rows) -> typing.NoReturn:
        raise ValueError("openpyxl目前不支持选中多行")


    def select_columns(self, columns) -> typing.NoReturn:
        raise ValueError("openpyxl目前不支持选中多列")


    def get_name(self) -> str:
        return self.worksheet.title

    
    def append_column(self, values, begin_row_index = 1) -> typing.NoReturn:
        if not isinstance(values, (list, tuple)):
            values = [values]

        column_index = self.get_column_count() + 1
        for row_index, val in enumerate(values):
           self.worksheet.cell(row=begin_row_index + row_index, column=column_index).value = val


    def insert_column(self, column_name, values, begin_row_index = 1) -> typing.NoReturn:
        if not isinstance(values, (list, tuple)):
            values = [values]

        column_index = openpyxl.utils.cell.column_index_from_string(column_name)
        self.worksheet.insert_cols(column_index)
        for row_index, val in enumerate(values):
           self.worksheet.cell(row=begin_row_index + row_index, column=column_index).value = val


    def get_first_free_column(self) -> str:
        column_index = self.get_column_count() + 1
        return openpyxl.utils.cell.get_column_letter(column_index)


    def get_first_free_row(self) -> int:
        return self.get_row_count() + 1


    def remove_column(self, column_name) ->typing.NoReturn:
        column_index = openpyxl.utils.cell.column_index_from_string(column_name)
        self.worksheet.delete_cols(column_index)

    
    def get_first_free_row_on_column(self, column_name) -> int:
        values = self.get_column(column_name)

        #处理[元素, 元素, 元素, None, None, None]的情况
        for index in range(len(values) - 1, -1, -1):
            if values[index] != None:
                return index + 2
        #空列
        return 1

    
    def clear_range(self, begin_row_num, begin_column_name, end_row_num, end_column_name) -> typing.NoReturn:
        raise ValueError("openpyxl目前不支持区域清空")

        #https://stackoverflow.com/questions/36582460/how-to-clear-a-range-of-values-in-an-excel-workbook-using-openpyxl/52646967
        #将区域单元格设置成None之后，但不能正确获取当前的有效行数
        #openpyxl不能像office一样动态调整sheet页的已用范围
        
        #begin_column_index = openpyxl.utils.cell.column_index_from_string(begin_column_name)
        #end_column_index = openpyxl.utils.cell.column_index_from_string(end_column_name)

        #for i in range(begin_column_index, end_column_index+1):
        #    for j in range(begin_row_num, end_row_num + 1):
        #        self.worksheet.cell(row=j, column=i).value = None


    def _get_range_cells(self, begin_row_num, begin_column_name, end_row_num, end_column_name):
        cells = []
        for row in self.worksheet[f'{begin_column_name}{begin_row_num}':f'{end_column_name}{end_row_num}']:
            cells.append([(cell.data_type, cell.value) for cell in row])
        return cells


    def copy_range(self, begin_row_num, begin_column_name, end_row_num, end_column_name) -> typing.NoReturn:
        OpenPyxlWorkSheet.copyed_cells = self._get_range_cells(begin_row_num, begin_column_name, end_row_num, end_column_name)


    def copy_rows(self, rows) -> typing.NoReturn:
        OpenPyxlWorkSheet.copyed_cells = []

        column_count = self.get_column_count()
        if column_count == 0:
            return
        end_column_name = openpyxl.utils.cell.get_column_letter(column_count)

        for row in rows:
            OpenPyxlWorkSheet.copyed_cells += self._get_range_cells(row, 'A', row, end_column_name)


    def copy_columns(self, columns) -> typing.NoReturn:
        OpenPyxlWorkSheet.copyed_cells = []

        row_count = self.get_row_count()
        if row_count == 0:
            return

        cells = []
        for column in columns:
            cells.append(self._get_range_cells(1, column, row_count, column))

        for i in range(0, row_count):
            row_value = []
            for j in range(0, len(columns)):
                row_value += cells[j][i]
            OpenPyxlWorkSheet.copyed_cells.append(row_value)


    def paste_range(self, row_num, column_name, copy_formula=True) -> typing.NoReturn:
        values = []
        for row_cell in OpenPyxlWorkSheet.copyed_cells:
            values.append([cell[1] if copy_formula or cell[0] != 'f' else None for cell in row_cell])
        self.set_range(row_num, column_name, values)