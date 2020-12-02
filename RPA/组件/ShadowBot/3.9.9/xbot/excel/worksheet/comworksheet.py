import typing
import openpyxl

from .baseworksheet import BaseWorkSheet

class ComWorkSheet(BaseWorkSheet):
    def __init__(self, worksheet):
        self.worksheet = worksheet


    def get_cell(self, row_num, col_name) -> any:
        return self.worksheet.Range(f"{col_name}{row_num}").Value


    def get_row(self, row_num) -> typing.List[any]:
        if self.get_row_count() == 0:
            return []

        used_range_row_begin = self.worksheet.UsedRange.Row
        used_range_row_end = self.worksheet.UsedRange.Row + self.worksheet.UsedRange.Rows.Count - 1
        
        if row_num < used_range_row_begin or row_num > used_range_row_end:
            return []
        else:
            used_range_column_end = self.worksheet.UsedRange.Column + self.worksheet.UsedRange.Columns.Count - 1
            tuple_result = self.worksheet.Range(self.worksheet.Cells(row_num, 1), self.worksheet.Cells(row_num, used_range_column_end)).Value
            #当只有 A1 单元格有数据时，会直接返回 A1 单元格的值，而非列表
            if isinstance(tuple_result, (tuple)):
                return list(tuple_result[0])
            else:
                return [tuple_result]


    def get_column(self, col_name) -> typing.List[any]:        
        end_row_num = self.get_row_count()
        if end_row_num == 0:
            return []
        
        begin_row_num = 1
        # Range的效率和其实际包含的数据无关，和调用Range的次数有关，所以优化成下面这种方式提高效率
        tuple_result = self.worksheet.Range(f"{col_name}{begin_row_num}:{col_name}{end_row_num}").Value
        if isinstance(tuple_result, (tuple)):
            return [row[0] for row in tuple_result]
        else:
            return [tuple_result]
        

    def get_range(self, begin_row_num, begin_column_name, end_row_num, end_column_name) -> typing.List[typing.List[any]]:
        tuple_result = self.worksheet.Range(f"{begin_column_name}{begin_row_num}:{end_column_name}{end_row_num}").Value
        if not isinstance(tuple_result, tuple):
            tuple_result = [[tuple_result]]
        return [list(item) for item in tuple_result]


    def set_cell(self, row_num, col_name, value) -> typing.NoReturn:
        self.worksheet.Range(f"{col_name}{row_num}").Value = [[value]]


    def set_row(self, row_num, values, begin_column_name='A') -> typing.NoReturn:
        # 2.1 将单个值转成[]
        if not isinstance(values, (list, tuple)):
            values = [values]

        # 2.2 计算插入点并插入内容
        begin_column_index = openpyxl.utils.cell.column_index_from_string(begin_column_name)
        begin_cell = self.worksheet.Cells(row_num, begin_column_index)
        end_cell = self.worksheet.Cells(row_num, begin_column_index + len(values) - 1)

        self.worksheet.Range(begin_cell, end_cell).Value = [values]  # 通过sheet.Range的方式写入内容时，com需要二维结构的数据
    

    def set_column(self, col_name, values, begin_row_num = 1) -> typing.NoReturn:
        # 2.1 将单个值转成[]
        if not isinstance(values, (list, tuple)):    #单个值的情况
            values = [values]
        # 2.2 计算插入点
        col_name_index = openpyxl.utils.cell.column_index_from_string(col_name)
        begin_cell = self.worksheet.Cells(begin_row_num, col_name_index)
        end_cell = self.worksheet.Cells(begin_row_num + len(values) - 1, col_name_index)
        # 2.3 插入内容
        self.worksheet.Range(begin_cell, end_cell).Value = [[item] for item in values]  # 通过sheet.Range的方式输入内容时，com需要二维结构的数据


    def set_range(self, row_num, col_name, values) -> typing.NoReturn:
        #2.1 转成[[]]
        if not isinstance(values, (list, tuple)):    #单个值的情况 ()
            values = [[values]]
        else:
            if len(values) == 0:           # []的情况
                values = [[]]
            else:
                if not isinstance(values[0], (list, tuple)):   # [1,2,3,4,5]的情况 (一维)
                    values = [values]
        #2.2 写入数据
        column_index = openpyxl.utils.cell.column_index_from_string(col_name)
        for row_value in values:
            for col_index, col_val in enumerate(row_value, start = column_index):
                self.worksheet.Cells(row_num, col_index).value = col_val
            row_num += 1


    def append_row(self, values, begin_column_name = 'A') -> typing.NoReturn:
        # 2.1 将单个值转成[]
        if not isinstance(values, (list, tuple)):
            values = [values]

        # 2.2 计算插入点并插入内容
        row_num = self.get_row_count() + 1
        begin_column_index = openpyxl.utils.cell.column_index_from_string(begin_column_name)
        begin_cell = self.worksheet.Cells(row_num, begin_column_index)
        end_cell = self.worksheet.Cells(row_num, begin_column_index + len(values) - 1)

        self.worksheet.Range(begin_cell, end_cell).Value = [values]  # 通过sheet.Range的方式写入内容时，com需要二维结构的数据


    def insert_row(self, row_num, values, begin_column_name = 'A') -> typing.NoReturn:
        # 2.1 将单个值转成[]
        if not isinstance(values, (list, tuple)):
            values = [values]
        
        # 2.2 计算插入点并插入内容
        self.worksheet.Range(f'{row_num}:{row_num}').Insert()
        begin_column_index = openpyxl.utils.cell.column_index_from_string(begin_column_name)
        begin_cell = self.worksheet.Cells(row_num, begin_column_index)
        end_cell = self.worksheet.Cells(row_num, begin_column_index + len(values) - 1)

        self.worksheet.Range(begin_cell, end_cell).Value = [values]  # 通过sheet.Range的方式写入内容时，com需要二维结构的数据


    def remove_row(self, row_num) -> typing.NoReturn:
        self.worksheet.Range(f'{row_num}:{row_num}').Delete()


    def clear(self) -> typing.NoReturn:
        total_rows_count = self.get_row_count()
        if total_rows_count > 0:
            self.worksheet.Range(f'1:{total_rows_count}').Delete()
    
    
    def get_row_count(self) -> int:
        if self.worksheet.UsedRange.Row == 1 and self.worksheet.UsedRange.Column == 1 and self.worksheet.UsedRange.Rows.Count == 1 and self.worksheet.UsedRange.Columns.Count == 1 and self.worksheet.Cells(1,1).value == None:
            return 0
        else:
            return self.worksheet.UsedRange.Rows.Count + self.worksheet.UsedRange.Row - 1

    def get_column_count(self) -> int:
        if self.worksheet.UsedRange.Row == 1 and self.worksheet.UsedRange.Column == 1 and self.worksheet.UsedRange.Rows.Count == 1 and self.worksheet.UsedRange.Columns.Count == 1 and self.worksheet.Cells(1,1).value == None:
            return 0
        else:
            return self.worksheet.UsedRange.Columns.Count + self.worksheet.UsedRange.Column - 1


    def select_range(self, begin_row_num, begin_column_name, end_row_num, end_column_name) -> typing.NoReturn:
        self.worksheet.Range(f"{begin_column_name}{begin_row_num}:{end_column_name}{end_row_num}").Select()


    def select_rows(self, rows) -> typing.NoReturn:
        rows = [str(row)+':'+str(row) for row in rows]
        self.worksheet.Range(','.join(rows)).Select()


    def select_columns(self, columns) -> typing.NoReturn:
        columns = [column+':'+column for column in columns]
        self.worksheet.Range(','.join(columns)).Select()


    def get_name(self) -> str:
        return self.worksheet.name

    
    def append_column(self, values, begin_row_index = 1) -> typing.NoReturn:
        column_index = self.get_column_count() + 1
        column_name = openpyxl.utils.cell.get_column_letter(column_index)
        self.insert_column(column_name, values, begin_row_index)


    def insert_column(self, column_name, values, begin_row_index = 1) -> typing.NoReturn:
        if not isinstance(values, (list, tuple)):
            values = [values]
        
        #计算插入点并插入内容
        self.worksheet.Range(f'{column_name}:{column_name}').Insert()
        column_index = openpyxl.utils.cell.column_index_from_string(column_name)
        begin_cell = self.worksheet.Cells(begin_row_index, column_index)
        end_cell = self.worksheet.Cells(begin_row_index + len(values) - 1, column_index)

        self.worksheet.Range(begin_cell, end_cell).Value = [[value] for value in values]  # 通过sheet.Range的方式写入内容时，com需要二维结构的数据

    
    def get_first_free_column(self) -> str:
        column_index = self.get_column_count() + 1
        return openpyxl.utils.cell.get_column_letter(column_index)


    def get_first_free_row(self) -> int:
        return self.get_row_count() + 1


    def remove_column(self, column_name) ->typing.NoReturn:
        self.worksheet.Range(f'{column_name}:{column_name}').Delete()


    def get_first_free_row_on_column(self, column_name) -> int:
        values = self.get_column(column_name)

        #处理[元素, 元素, 元素, None, None, None]的情况
        for index in range(len(values) - 1, -1, -1):
            if values[index] != None:
                return index + 2
        #空列
        return 1

    def clear_range(self, begin_row_num, begin_column_name, end_row_num, end_column_name) -> typing.NoReturn:
        self.worksheet.Range(f'{begin_column_name}{begin_row_num}:{end_column_name}{end_row_num}').Clear()


    def copy_range(self, begin_row_num, begin_column_name, end_row_num, end_column_name) -> typing.NoReturn:
        self.worksheet.Range(f'{begin_column_name}{begin_row_num}:{end_column_name}{end_row_num}').Copy()


    def copy_rows(self, rows) -> typing.NoReturn:
        end_column_name = openpyxl.utils.cell.get_column_letter(self.get_column_count())
        rows = [f'A{row}:{end_column_name}{row}' for row in rows]
        self.worksheet.Range(','.join(rows)).Copy()


    def copy_columns(self, columns) -> typing.NoReturn:
        end_row_num = self.get_row_count()
        columns = [f'{column}1:{column}{end_row_num}' for column in columns]
        self.worksheet.Range(','.join(columns)).Copy()


    def paste_range(self, row_num, column_name, copy_formula=True) -> typing.NoReturn:
        self.worksheet.Range(f'{column_name}{row_num}').PasteSpecial(-4104 if copy_formula else -4163)